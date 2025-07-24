import os
import openpyxl
from GlobalServices import Log, Notify
from XLSXExceptions import XLSXSheetNotFoundError

class ExcelService:
    def __init__(self, file_path: str):
        file_path = os.getenv('mainPath') + file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.strip_sheet_names_whitespace()

    def get_sheet(self, sheet_name: str):
        if sheet_name not in self.workbook.sheetnames:
            # raise XLSXSheetNotFoundError(f"Sheet '{sheet_name}' not found")
            Log.add('info', 'ExcelService', f"Arkusz '{sheet_name}' nie został znaleziony")
            return None

        return self.workbook[sheet_name]

    def strip_sheet_names_whitespace(self):
        """Usuwa spacje z końca nazw arkuszy."""
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            new_sheet_name = sheet_name.strip().lower()
            if new_sheet_name != sheet_name:
                sheet.title = new_sheet_name
                Log.add('info', 'ExcelService', f"Zmieniono nazwę arkusza z '{sheet_name}' nа '{new_sheet_name}'")

    def find_user_in_sheet(self, sheet, user: str, column: int, max_row: int):
        """Wyszukuje użytkownika w arkuszu"""
        result = []
        for row in range(1, max_row + 1):
            cell_value = sheet.cell(row=row, column=column).value
            cell_value = cell_value.strip() if cell_value else ''
            if cell_value == user:
                result.append((cell_value, column, row))
        return result